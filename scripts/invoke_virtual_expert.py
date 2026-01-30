#!/usr/bin/env python3
"""
Invoke Virtual Expert - Generate evaluation report in Evaluation Form format.

Usage:
    python invoke_virtual_expert.py --findings findings.json --product-name "Product Name" --evaluator "Your Name" --output report.xlsx
"""

import argparse
import json
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    exit(1)


def consolidate_findings_by_step(findings):
    """Consolidate findings by step, limiting to 3 heuristics per step."""
    step_findings = {}
    
    for finding in findings:
        step = finding.get('step', 999)
        location = finding.get('location', '')
        
        if step not in step_findings:
            step_findings[step] = {
                'step': step,
                'location': location,
                'heuristics': [],
                'issues': [],
                'recommendations': [],
                'severity': 'Low'
            }
        
        # Add heuristic (limit to 3)
        heuristic = finding.get('heuristic', '')
        if heuristic and heuristic not in step_findings[step]['heuristics']:
            if len(step_findings[step]['heuristics']) < 3:
                step_findings[step]['heuristics'].append(heuristic)
        
        # Add issue
        issue = finding.get('description', '')
        if issue:
            step_findings[step]['issues'].append(issue)
        
        # Add recommendation
        recommendation = finding.get('recommendation', '')
        if recommendation:
            step_findings[step]['recommendations'].append(recommendation)
        
        # Update severity (keep highest)
        severity_order = {'Critical': 4, 'High': 3, 'Medium': 2, 'Low': 1}
        current_severity = finding.get('severity', 'Low')
        if severity_order.get(current_severity, 0) > severity_order.get(step_findings[step]['severity'], 0):
            step_findings[step]['severity'] = current_severity
    
    return sorted(step_findings.values(), key=lambda x: x['step'])


def create_evaluation_form(wb, findings):
    """Create Evaluation Form sheet with consolidated findings."""
    ws = wb.create_sheet("Evaluation Form", 0)
    
    headers = ["Task", "Violated Heuristic", "Issue", "Recommendation", "Severity"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    
    severity_validation = DataValidation(type="list", formula1='"Critical,High,Medium,Low"')
    ws.add_data_validation(severity_validation)
    
    # Consolidate findings by step
    consolidated = consolidate_findings_by_step(findings)
    
    # Add consolidated findings
    row_idx = 2
    for finding in consolidated:
        # Task (Bold font)
        step = finding['step']
        location = finding['location']
        task_text = f"Step {step}: {location}" if step != 999 else location
        
        task_cell = ws.cell(row_idx, 1, task_text)
        task_cell.font = Font(bold=True)
        task_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Violated Heuristic (multiple lines, max 3)
        heuristics_text = ',\n'.join(finding['heuristics'][:3])
        heuristic_cell = ws.cell(row_idx, 2, heuristics_text)
        heuristic_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Issue (concise, multiple issues separated by semicolon)
        issues_text = '；'.join(finding['issues'])
        issue_cell = ws.cell(row_idx, 3, issues_text)
        issue_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Recommendation (concise, multiple recommendations separated by semicolon)
        recommendations_text = '；'.join(finding['recommendations'])
        recommendation_cell = ws.cell(row_idx, 4, recommendations_text)
        recommendation_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Severity
        severity = finding['severity']
        severity_cell = ws.cell(row_idx, 5, severity)
        severity_cell.alignment = Alignment(horizontal='center', vertical='top')
        
        # Apply validation
        severity_validation.add(severity_cell)
        
        # Color coding for severity
        if severity == "Critical":
            severity_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            severity_cell.font = Font(bold=True)
        elif severity == "High":
            severity_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            severity_cell.font = Font(bold=True)
        elif severity == "Medium":
            severity_cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
            severity_cell.font = Font(bold=True)
        else:  # Low
            severity_cell.fill = PatternFill(start_color="95D5B2", end_color="95D5B2", fill_type="solid")
            severity_cell.font = Font(bold=True)
        
        row_idx += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 25  # Task
    ws.column_dimensions['B'].width = 35  # Violated Heuristic
    ws.column_dimensions['C'].width = 45  # Issue
    ws.column_dimensions['D'].width = 45  # Recommendation
    ws.column_dimensions['E'].width = 12  # Severity
    
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:E{row_idx - 1}"


def main():
    parser = argparse.ArgumentParser(description='Invoke Virtual Expert - Generate evaluation report')
    parser.add_argument('--findings', required=True, help='Path to findings JSON file')
    parser.add_argument('--product-name', required=True, help='Product name')
    parser.add_argument('--evaluator', required=True, help='Evaluator name')
    parser.add_argument('--output', default='virtual-expert-report.xlsx', help='Output Excel file')
    
    args = parser.parse_args()
    
    with open(args.findings, 'r', encoding='utf-8') as f:
        findings = json.load(f)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    create_evaluation_form(wb, findings)
    
    output_path = Path(args.output)
    wb.save(output_path)
    
    print(f"✓ Virtual Expert report generated: {output_path.absolute()}")
    print(f"  Total steps evaluated: {len(set(f.get('step', 999) for f in findings))}")
    print(f"  Product: {args.product_name}")
    print(f"  Evaluated by: {args.evaluator}")


if __name__ == '__main__':
    main()
